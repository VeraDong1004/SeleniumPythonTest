# coding=utf-8
import os
import stat

from openpyxl import load_workbook  # 处理 .xlsx
import xlrd  # 处理 .xls
from xlutils.copy import copy



class ExcelUtil:
    def __init__(self, excel_path=None, index=0):
        # 默认路径处理（建议使用相对路径）
        if excel_path is None:
            self.excel_path = os.path.join(os.getcwd(), 'config', 'casedata.xlsx')
        else:
            self.excel_path=excel_path
        self.index = index
        self.data = None
        self.table = None
        self.rows = 0

        # 根据文件格式选择读取方式
        self.load_excel()

    def load_excel(self):
        """根据文件后缀选择不同的库加载Excel"""
        if self.excel_path.endswith('.xlsx'):
            # 使用 openpyxl 读取 .xlsx
            wb = load_workbook(self.excel_path)
            sheets = wb.sheetnames
            self.data = wb  # 保存 workbook 对象
            self.table = wb[sheets[self.index]]
            self.rows = self.table.max_row

        else:
            # 使用 xlrd 读取 .xls
            self.data = xlrd.open_workbook(self.excel_path)
            self.table = self.data.sheet_by_index(self.index)
            self.rows = self.table.nrows
        return self.data,self.table,self.rows

    def get_data(self):
        """获取Excel数据，返回二维列表"""
        result = []
        if self.excel_path.endswith('.xlsx'):
            # 处理 .xlsx
            for row in self.table.iter_rows(values_only=True):
                result.append(list(row))
        else:
            # 处理 .xls
            rows=self.get_lines()
            if rows!=None:
                for i in range(rows):
                    result.append(self.table.row_values(i))
            return None
        return result

    def get_lines(self):
        '''获取excel行数'''
        if self.excel_path.endswith('.xlsx'):
            rows= self.table.max_row
        else:
            rows=self.table.nrows
            if rows>=1:
                return rows
            return None

    def get_col_value(self,row,col):
        '''获取单元格的数据:
            float->str
            去除.0后显示'''
        if self.get_lines()>row:
            data=self.table.cell(row,col).value
            str_data=str(data)
            fin_data = str_data.rstrip('0').rstrip('.') if '.' in str_data else str_data
            return fin_data
        return None

    def write_value(self,row,col,value):
        '''写入数据'''
        if self.excel_path.endswith('.xlsx'):
            wb=load_workbook(self.excel_path)
            ws=wb['Sheet1']#按表名选择工作表
            ws.cell(row,col,value)#写入数据
            wb.save(self.excel_path)#保存<会覆盖原文件>

        else:
            write_data=copy(xlrd.open_workbook(self.excel_path))
            #选择sheet0
            sheet=write_data.get_sheet(0)
            #写入数据
            sheet.write(row,col,value)
            #保存文件
            write_data.save(self.excel_path)


if __name__ == '__main__':
    # 示例用法
    #ex = ExcelUtil()
    ex=ExcelUtil('E:\\00-PycharmProjects\\SeleniumPythonEx\\config\\keywords.xls')# 默认读取 casedata.xlsx
    #ex.write_value(15,1,'test')
    print(ex.get_col_value(3,7))